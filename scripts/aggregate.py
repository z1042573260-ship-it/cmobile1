"""
多源数据汇总层
------------
将所有爬虫结果汇集、去重，输出统一 JSON，供 AI 分析使用。

支持三种输入源混用：
  1. 爬虫 JSON — spider.run() 的直接输出
  2. 审核表 Excel — review_data.py --export-excel 的人工审核结果
  3. 数据库 — raw_projects 表（需 Flask app context）

用法：
  python scripts/aggregate.py --from-json "data/spider_test/yantai_districts_merged.json"
  python scripts/aggregate.py --from-excel "data/spider_test/*审核表*.xlsx"
  python scripts/aggregate.py --from-json "data/spider_test/*.json" --from-excel "data/*审核表*.xlsx"
  python scripts/aggregate.py --from-json ... --export-json data/merged_for_ai.json
  python scripts/aggregate.py --from-json ... --export-excel data/merged_for_ai.xlsx
"""
from __future__ import annotations

import glob
import json
import os
import re
import sys
from difflib import SequenceMatcher
from pathlib import Path
from typing import Optional

import pandas as pd
from loguru import logger

# ---- 标准字段名（所有输入源映射到这套字段） ----
STANDARD_FIELDS = [
    "title", "content", "source_url", "source_name",
    "publish_date", "relevance_score", "score_detail",
    "district_extracted", "scale_extracted",
    "investment_extracted", "project_nature",
]

# 审核表 Excel 列名 → 标准字段名 映射
EXCEL_COLUMN_MAP = {
    "标题": "title",
    "区县": "district_extracted",
    "规模": "scale_extracted",
    "投资额": "investment_extracted",
    "项目性质": "project_nature",
    "发布日期": "publish_date",
    "URL": "source_url",
    "评分": "relevance_score",
    "质量": "quality",
    "内容摘要(前200字)": "content",
    "内容摘要": "content",
    "审核结果": "review_result",
}


def collect_from_json(path_or_pattern: str) -> list[dict]:
    """
    从爬虫 JSON 文件收集结果。

    支持：
      - 单个文件: "data/spider_test/xxx.json"
      - glob 通配符: "data/spider_test/yantai_*.json"

    自动识别两种 JSON 格式：
      格式A: [{"title": ..., "content": ..., ...}, ...]  (spider 直接输出)
      格式B: {"results": [...], "stats": {...}}           (test_spider.py 包装)
    """
    paths = glob.glob(path_or_pattern, recursive=True)
    if not paths:
        logger.warning(f"未匹配到 JSON 文件: {path_or_pattern}")
        return []

    all_records = []
    for path in paths:
        logger.info(f"读取 JSON: {path}")
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception as e:
            logger.error(f"读取失败 {path}: {e}")
            continue

        # 格式B: {"results": [...]}
        if isinstance(data, dict):
            results = data.get("results", data.get("items", []))
            if results:
                data = results
            elif "title" in data:
                # 单个记录
                data = [data]
            else:
                logger.warning(f"无法识别的 JSON 格式: {path}")
                continue

        if not isinstance(data, list):
            logger.warning(f"JSON 内容不是列表: {path}")
            continue

        # 规范化每条记录
        for item in data:
            if not isinstance(item, dict):
                continue
            if not item.get("title"):
                continue
            normalized = _normalize_record(item, source_file=os.path.basename(path))
            all_records.append(normalized)

    logger.info(f"从 {len(paths)} 个 JSON 文件收集到 {len(all_records)} 条记录")
    return all_records


def collect_from_excel(path_or_pattern: str) -> list[dict]:
    """
    从审核表 Excel 收集结果。

    读取 review_data.py --export-excel 生成的审核表，
    列名：序号/评分/质量/标题/区县/规模/投资额/项目性质/发布日期/URL/内容摘要(前200字)/审核结果

    只收集有内容的行（跳过空行和表头）。
    """
    paths = glob.glob(path_or_pattern, recursive=True)
    if not paths:
        logger.warning(f"未匹配到 Excel 文件: {path_or_pattern}")
        return []

    all_records = []
    for path in paths:
        logger.info(f"读取 Excel: {path}")
        try:
            df = pd.read_excel(path)
        except Exception as e:
            logger.error(f"读取失败 {path}: {e}")
            continue

        # 跳过"使用说明" sheet — pandas 默认读第一个 sheet，审核表只有一个数据 sheet
        # 如果第一个 sheet 是使用说明，跳过取第二个
        if "使用说明" in str(df.iloc[0, 0]) if len(df) > 0 else False:
            try:
                xl = pd.ExcelFile(path)
                data_sheets = [s for s in xl.sheet_names if "使用说明" not in s]
                if data_sheets:
                    df = pd.read_excel(path, sheet_name=data_sheets[0])
            except Exception:
                pass

        for _, row in df.iterrows():
            record = _excel_row_to_record(row)
            if record:
                record["_source_file"] = os.path.basename(path)
                all_records.append(record)

    logger.info(f"从 {len(paths)} 个 Excel 文件收集到 {len(all_records)} 条记录")
    return all_records


def collect_from_db(app, limit: int = 500) -> list[dict]:
    """
    从数据库 raw_projects 表收集未处理数据。
    需要 Flask app context。
    """
    from database.db_manager import get_unprocessed_raws

    with app.app_context():
        raw_list = get_unprocessed_raws(limit=limit)

    records = []
    for raw in raw_list:
        records.append({
            "title": raw.title,
            "content": raw.content or "",
            "source_url": raw.source_url or "",
            "source_name": raw.source_name or "",
            "publish_date": raw.publish_date or "",
            "relevance_score": 0,
            "score_detail": {},
            "district_extracted": "",
            "scale_extracted": "",
            "investment_extracted": "",
            "project_nature": "",
            "_source_db_id": raw.id,
        })

    logger.info(f"从数据库收集到 {len(records)} 条未处理记录")
    return records


def deduplicate(records: list[dict], threshold: float = 0.85) -> list[dict]:
    """
    标题相似度去重（按区县分组，组内去重）。

    核心逻辑：不同区县的项目不可能重复，即使标题相似。
    比如海阳市和龙口市都发"建设工程规划许可证核发批后公布"，
    但它们是不同区县的不同项目，不应去重。

    去重规则：
      1. 先按 district_extracted 分组
      2. 同一区县内，标题相似度 > threshold 的视为重复
      3. 跨区县不去重
      4. 区县为空的归入"未知"组，组内去重
    """
    if not records:
        return []

    # 按区县分组
    groups: dict[str, list[int]] = {}  # district → [index in original records]
    for i, rec in enumerate(records):
        district = rec.get("district_extracted", "") or rec.get("district_name", "") or rec.get("district", "")
        district = str(district).strip()
        if not district:
            district = "__未知区县__"
        groups.setdefault(district, []).append(i)

    all_remove = set()
    district_stats = {}

    for district, indices in groups.items():
        if len(indices) <= 1:
            continue

        # 组内去重
        # 按标题长度降序排列（长的优先保留）
        indexed = [(i, records[i]) for i in indices]
        indexed.sort(key=lambda x: len(x[1].get("title", "")), reverse=True)

        n = len(indexed)
        to_remove = set()
        titles = [r.get("title", "") for _, r in indexed]

        for i in range(n):
            if i in to_remove:
                continue
            for j in range(i + 1, n):
                if j in to_remove:
                    continue
                ratio = SequenceMatcher(None, titles[i], titles[j]).ratio()
                if ratio > threshold:
                    to_remove.add(j)

        # 转换回原始索引
        for j in to_remove:
            all_remove.add(indexed[j][0])

        removed_in_group = len(to_remove)
        if removed_in_group > 0:
            district_stats[district] = (len(indices), removed_in_group)

    result = [r for i, r in enumerate(records) if i not in all_remove]
    removed = len(records) - len(result)

    logger.info(f"去重: {len(records)} → {len(result)} 条（移除 {removed} 条重复，阈值={threshold}）")
    if district_stats:
        for district, (total, rm) in sorted(district_stats.items()):
            logger.debug(f"  {district}: {total} → {total - rm}（移除 {rm} 条）")

    return result


def export_json(records: list[dict], path: str) -> None:
    """导出为 JSON 文件。"""
    # 确保目录存在
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    logger.info(f"已导出 JSON: {path} ({len(records)} 条)")


def export_excel(records: list[dict], path: str) -> None:
    """导出为 Excel 文件（标准审核表格式）。"""
    Path(path).parent.mkdir(parents=True, exist_ok=True)

    rows = []
    for i, r in enumerate(records):
        score = r.get("relevance_score", 0)
        if score >= 5:
            quality = "高"
        elif score >= 3:
            quality = "中"
        elif score > 0:
            quality = "低"
        else:
            quality = "—"

        content = r.get("content", "")
        summary = content[:200] if content else ""

        rows.append({
            "序号": i + 1,
            "评分": score,
            "质量": quality,
            "标题": r.get("title", ""),
            "区县": r.get("district_extracted", ""),
            "规模": r.get("scale_extracted", ""),
            "投资额": r.get("investment_extracted", ""),
            "项目性质": r.get("project_nature", ""),
            "发布日期": str(r.get("publish_date", "")),
            "URL": r.get("source_url", ""),
            "内容摘要(前200字)": summary,
            "审核结果": "",
        })

    df = pd.DataFrame(rows)

    # 样式写入（与 review_data.py 一致的格式）
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总数据"

    # 表头
    headers = list(rows[0].keys()) if rows else []
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 数据行
    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mid_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    data_font = Font(name="微软雅黑", size=10)
    url_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

    for row_idx, row_data in enumerate(rows):
        excel_row = row_idx + 2
        quality = row_data["质量"]

        if quality == "高":
            fill = high_fill
        elif quality == "中":
            fill = mid_fill
        elif quality == "低":
            fill = low_fill
        else:
            fill = PatternFill()

        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = url_font if header == "URL" else data_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="center", wrap_text=(header == "内容摘要(前200字)"))
            cell.border = thin_border

    # 列宽
    col_widths = {
        "序号": 6, "评分": 6, "质量": 6, "标题": 45, "区县": 10,
        "规模": 18, "投资额": 14, "项目性质": 12, "发布日期": 14,
        "URL": 25, "内容摘要(前200字)": 40, "审核结果": 12,
    }
    for col_idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(header, 12)

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # 使用说明 sheet
    ws2 = wb.create_sheet("使用说明")
    ws2["A1"] = "数据汇总说明"
    ws2["A1"].font = Font(name="微软雅黑", bold=True, size=14)
    ws2["A3"] = f"本文件由 scripts/aggregate.py 自动生成"
    ws2["A4"] = f"记录总数: {len(records)} 条"
    ws2["A5"] = "质量分类: 高(评分≥5) / 中(评分3-4) / 低(评分1-2) / —(未评分)"
    ws2["A6"] = "颜色: 绿=高质量 黄=中等 红=低质量"
    ws2["A7"] = ""
    ws2["A7"] = "下一步: 将此文件用于 AI 分析"
    ws2["A8"] = "  python scripts/aggregate.py --from-excel <本文件> --export-json data/for_ai.json"
    ws2["A9"] = "  python -c \"from processor.business_analyzer import run_business_pipeline; ...\""

    wb.save(path)
    logger.info(f"已导出 Excel: {path} ({len(records)} 条)")


def _normalize_record(item: dict, source_file: str = "") -> dict:
    """将爬虫 JSON 记录规范化为标准字段。"""
    record = {}
    for field in STANDARD_FIELDS:
        record[field] = item.get(field, "")

    # 补充字段
    if not record["source_name"]:
        record["source_name"] = item.get("district_name", item.get("source", ""))
    record["_source_file"] = source_file

    # 保留爬虫特有字段（用于后续分析）
    for extra_key in ["district", "district_name", "project_name",
                       "project_location", "construction_unit",
                       "total_investment", "cooperation_mode",
                       "lead_unit", "land_use_type"]:
        if extra_key in item and item[extra_key]:
            record[extra_key] = item[extra_key]

    return record


def _excel_row_to_record(row: pd.Series) -> Optional[dict]:
    """将审核表 Excel 行转换为标准字段记录。"""
    title = str(row.get("标题", "")).strip()
    if not title or title in ("nan", "标题", "None"):
        return None

    record = {}
    for excel_col, std_col in EXCEL_COLUMN_MAP.items():
        val = row.get(excel_col, "")
        if pd.isna(val):
            record[std_col] = ""
        else:
            record[std_col] = str(val).strip()

    # 评分转 int
    try:
        record["relevance_score"] = int(float(record.get("relevance_score", 0)))
    except (ValueError, TypeError):
        record["relevance_score"] = 0

    # 确保必填字段存在
    record.setdefault("title", title)
    record.setdefault("content", "")
    record.setdefault("source_url", "")
    record.setdefault("source_name", "")
    record.setdefault("publish_date", "")
    record.setdefault("score_detail", {})
    record.setdefault("project_nature", "")

    return record


# ==============================================================================
# CLI
# ==============================================================================

def main():
    import argparse

    parser = argparse.ArgumentParser(
        description="多源数据汇总 — 收集爬虫JSON + 审核表Excel → 去重 → 统一输出",
    )
    parser.add_argument(
        "--from-json", action="append", default=[],
        help="爬虫 JSON 文件（支持 glob），可多次指定",
    )
    parser.add_argument(
        "--from-excel", action="append", default=[],
        help="审核表 Excel 文件（支持 glob），可多次指定",
    )
    parser.add_argument(
        "--from-db", action="store_true",
        help="从数据库 raw_projects 表收集",
    )
    parser.add_argument(
        "--db-limit", type=int, default=500,
        help="数据库读取上限（默认500）",
    )
    parser.add_argument(
        "--dedup-threshold", type=float, default=0.85,
        help="标题相似度去重阈值（0-1，默认0.85）",
    )
    parser.add_argument(
        "--dedup", action="store_true",
        help="启用标题相似度去重（默认不去重，数据来源多样，交给 AI 处理）",
    )
    parser.add_argument(
        "--min-score", type=int, default=0,
        help="最低评分过滤（默认0=不过滤）",
    )
    parser.add_argument(
        "--export-json", type=str, default=None,
        help="导出 JSON 文件路径",
    )
    parser.add_argument(
        "--export-excel", type=str, default=None,
        help="导出 Excel 文件路径",
    )
    parser.add_argument(
        "--stats", action="store_true",
        help="只打印统计信息，不导出",
    )

    args = parser.parse_args()

    # 如果没有指定任何输入源，默认搜索 data/spider_test/ 下的 JSON
    if not args.from_json and not args.from_excel and not args.from_db:
        default_pattern = "data/spider_test/*.json"
        if glob.glob(default_pattern):
            logger.info(f"未指定输入源，默认使用: {default_pattern}")
            args.from_json = [default_pattern]
        else:
            logger.error("未找到任何输入数据，请用 --from-json / --from-excel / --from-db 指定")
            sys.exit(1)

    # ---- 收集 ----
    all_records = []

    for pattern in args.from_json:
        all_records.extend(collect_from_json(pattern))

    for pattern in args.from_excel:
        all_records.extend(collect_from_excel(pattern))

    if args.from_db:
        from flask import Flask
        from config.settings import DATABASE_URL, FLASK_SECRET_KEY
        from database.models import db

        app = Flask(__name__)
        app.config["SECRET_KEY"] = FLASK_SECRET_KEY
        app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
        app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
        db.init_app(app)
        all_records.extend(collect_from_db(app, limit=args.db_limit))

    if not all_records:
        logger.error("未收集到任何记录")
        sys.exit(1)

    logger.info(f"共收集 {len(all_records)} 条原始记录")

    # ---- 去重（默认跳过，数据来源多样，交给 AI 处理） ----
    if args.dedup:
        all_records = deduplicate(all_records, threshold=args.dedup_threshold)

    # ---- 评分过滤 ----
    if args.min_score > 0:
        before = len(all_records)
        all_records = [r for r in all_records if r.get("relevance_score", 0) >= args.min_score]
        logger.info(f"评分过滤 (≥{args.min_score}): {before} → {len(all_records)} 条")

    # ---- 统计 ----
    high = sum(1 for r in all_records if r.get("relevance_score", 0) >= 5)
    mid = sum(1 for r in all_records if 3 <= r.get("relevance_score", 0) < 5)
    low = sum(1 for r in all_records if 0 < r.get("relevance_score", 0) < 3)
    unrated = sum(1 for r in all_records if r.get("relevance_score", 0) == 0)

    logger.info(
        f"汇总结果: {len(all_records)} 条 "
        f"(高{high}/中{mid}/低{low}/未评分{unrated})"
    )

    if args.stats:
        # 打印来源分布
        sources = {}
        for r in all_records:
            src = r.get("source_name", r.get("_source_file", "未知"))
            sources[src] = sources.get(src, 0) + 1
        logger.info("来源分布:")
        for src, count in sorted(sources.items(), key=lambda x: -x[1]):
            logger.info(f"  {src}: {count} 条")
        return

    # ---- 导出 ----
    if args.export_json:
        export_json(all_records, args.export_json)

    if args.export_excel:
        export_excel(all_records, args.export_excel)

    if not args.export_json and not args.export_excel:
        logger.info("未指定导出目标（--export-json / --export-excel），只打印统计。")
        logger.info(f"最终 {len(all_records)} 条记录待使用")


if __name__ == "__main__":
    main()
