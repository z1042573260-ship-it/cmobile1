"""
MySQL 数据库 → 大屏 dashboard_data.json 导出
--------------------------------------------
从本地 MySQL（yantai_projects.projects）读取全部记录，
输出 frontend/data/dashboard_data.json，内容完整性原则：
  - project_name 完整（不截断；数据库已做唯一查重，"地块一/地块二"靠完整名区分）
  - investment/scale 保留数据库真实值（"待核实"等，不吞空）
  - map_points / project_list 附带数据库可展示字段（详情页直接用）

用法：
  python scripts/export_dashboard_db.py                # 导出到 frontend/data/dashboard_data.json
  python scripts/export_dashboard_db.py -o out.json    # 指定输出
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

try:
    import pymysql
except ImportError:
    print("[ERROR] 需要 pymysql：pip install pymysql")
    sys.exit(1)

try:
    from config.settings import DATABASE_URL
except Exception:
    DATABASE_URL = "mysql+pymysql://root:1234@localhost:3306/yantai_projects"


def parse_db_url(url: str) -> dict:
    """mysql+pymysql://user:pass@host:port/db?query → dict（支持 ssl_ca）"""
    core = url.split("://", 1)[1]
    cred, rest = core.split("@", 1)
    user, pwd = cred.split(":", 1)
    host_port, db = rest.split("/", 1)
    host, port = host_port.split(":", 1) if ":" in host_port else (host_port, "3306")
    cfg = {"user": user, "password": pwd, "host": host, "port": int(port), "database": db, "ssl": None}
    if "?" in db:
        db, query = db.split("?", 1)
        cfg["database"] = db
        for kv in query.split("&"):
            if "=" in kv:
                k, v = kv.split("=", 1)
                if k == "ssl_ca":
                    cfg["ssl"] = {"ca": v}
    return cfg


def _connect(cfg: dict):
    """按配置连接（TiDB 带 SSL）"""
    import pymysql as _pymysql
    kwargs = {k: cfg[k] for k in ("host", "port", "user", "password", "database", "ssl")}
    kwargs["charset"] = "utf8mb4"
    kwargs["cursorclass"] = _pymysql.cursors.DictCursor
    kwargs["connect_timeout"] = 30
    return _pymysql.connect(**kwargs)


def fetch_projects() -> list[dict]:
    cfg = parse_db_url(DATABASE_URL)
    conn = _connect(cfg)
    try:
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM projects ORDER BY priority DESC, id ASC")
            return list(cur.fetchall())
    finally:
        conn.close()


def clean(v):
    """None → ''，其余原样（不吞真实值如'待核实'）"""
    if v is None:
        return ""
    return v


def norm_date(v):
    """日期统一为横杠式（2026.02.06 → 2026-02-06）；非日期文本（'待核实'等）→ ''"""
    if not v:
        return ""
    s = str(v).replace(".", "-")
    if not re.match(r"^\d{4}-\d{1,2}-\d{1,2}", s):
        return ""
    return s


def stage_of(status):
    """status 完整描述 → 5 类阶段（大屏分组归档；DB 保留原文，详情页用 stage_detail）
    优先级：竣工 > 招标 > 施工 > 规划 > 待核实"""
    s = status or ""
    if any(k in s for k in ("竣工", "完工", "验收", "交付", "建成")):
        return "已竣工完工"
    if any(k in s for k in ("招标", "中标", "磋商", "资格预审", "开标")):
        return "招标阶段"
    if any(k in s for k in ("施工", "开工", "在建", "封顶", "主体")):
        return "施工阶段"
    if any(k in s for k in ("规划", "立项", "预审", "选址", "公示", "许可", "审批", "评估")):
        return "规划阶段"
    return "待核实"


def build_dashboard(rows: list[dict]) -> dict:
    total = len(rows)
    red = sum(1 for r in rows if "红色" in str(r.get("warning_level", "")))
    yellow = sum(1 for r in rows if "黄色" in str(r.get("warning_level", "")))
    valuable = sum(1 for r in rows if r.get("is_valuable"))
    districts = {r["district"] for r in rows if r.get("district") and r["district"] != "待核实"}

    # ---- 区县排名（统计口径：开发区→福山区、长岛→蓬莱区，与前端 statDistrict 一致）----
    dcount = {}
    for r in rows:
        d = r.get("district") or "未知"
        if d == "待核实":
            d = "未知"
        if d in ("开发区", "烟台开发区"):
            d = "福山区"
        elif d in ("长岛综合试验区", "长岛县", "长岛综试区"):
            d = "蓬莱区"
        dcount[d] = dcount.get(d, 0) + 1
    district_ranking = sorted(
        [{"name": k, "value": v} for k, v in dcount.items()],
        key=lambda x: x["value"], reverse=True)

    # ---- 预警/类型/阶段分布 ----
    warning_pie = [{"name": "红色预警", "value": red}, {"name": "黄色预警", "value": yellow},
                   {"name": "无预警", "value": total - red - yellow}]
    warning_pie = [x for x in warning_pie if x["value"] > 0]

    tcount = {}
    for r in rows:
        t = r.get("project_type") or "其他"
        tcount[t] = tcount.get(t, 0) + 1
    type_pie = sorted([{"name": k, "value": v} for k, v in tcount.items()],
                      key=lambda x: x["value"], reverse=True)

    scount = {}
    for r in rows:
        s = stage_of(r.get("status"))
        scount[s] = scount.get(s, 0) + 1
    # 固定 5 类顺序展示（规划/招标/施工/完工/待核实），0 值也保留
    stage_pie = [{"name": k, "value": scount.get(k, 0)}
                 for k in ("规划阶段", "招标阶段", "施工阶段", "已竣工完工", "待核实")]
    stage_pie = [x for x in stage_pie if x["value"] > 0]

    # ---- 时间趋势 ----
    # 日期统一横杠式；publish_date 为空 → start_date → end_date → created_at 兜底月份
    # （保证 timeline 覆盖全部记录，前端"项目总数"与 summary 一致）
    date_counts, red_dates, yellow_dates = {}, {}, {}
    for r in rows:
        d = norm_date(r.get("publish_date")) or norm_date(r.get("start_date")) \
            or norm_date(r.get("end_date")) or norm_date(r.get("created_at"))
        d = d[:7]
        if not d:
            continue
        date_counts[d] = date_counts.get(d, 0) + 1
        w = str(r.get("warning_level", ""))
        if "红色" in w:
            red_dates[d] = red_dates.get(d, 0) + 1
        elif "黄色" in w:
            yellow_dates[d] = yellow_dates.get(d, 0) + 1
    srt = lambda m: sorted([{"date": k, "value": v} for k, v in m.items()], key=lambda x: x["date"])
    timeline, red_timeline, yellow_timeline = srt(date_counts), srt(red_dates), srt(yellow_dates)

    # ---- map_points：红/黄预警 + 完整内容 ----
    map_points = []
    for r in rows:
        w = str(r.get("warning_level", ""))
        if "红色" in w:
            cat = "red"
        elif "黄色" in w:
            cat = "yellow"
        else:
            continue
        lng, lat = r.get("lng"), r.get("lat")
        if lng is None or lat is None:
            continue
        map_points.append({
            "name": clean(r.get("project_name")),            # 完整名，不截断
            "value": [float(lng), float(lat), int(r.get("priority") or 1)],
            "_lng": float(lng), "_lat": float(lat),          # 原始坐标（散开用，不入库）
            "district": clean(r.get("district")),
            "category": cat,
            "project_type": clean(r.get("project_type")),
            "stage": stage_of(r.get("status")),              # 5 类阶段（大屏分组/图钉过滤）
            "stage_detail": clean(r.get("status")),          # AI 完整阶段描述（详情页）
            "warning": w,
            # ---- 数据库可展示字段（详情页直接使用） ----
            "location": clean(r.get("location")),
            "investment": clean(r.get("investment")),
            "scale": clean(r.get("scale")),
            "content": clean(r.get("content")),
            "developer": clean(r.get("developer")),
            "contact_person": clean(r.get("contact_person")),
            "contact_phone": clean(r.get("contact_phone")),
            "publish_date": norm_date(r.get("publish_date")),
            "deadline": clean(r.get("deadline")),
            "start_date": clean(r.get("start_date")),
            "end_date": clean(r.get("end_date")),
            "need_base_station": clean(r.get("need_base_station")),
            "base_station_type": clean(r.get("base_station_type")),
            "coverage_area": clean(r.get("coverage_area")),
            "ai_reason": clean(r.get("ai_reason")),
            "ai_summary": clean(r.get("ai_summary")),
            "telecom_needs": clean(r.get("telecom_needs")),
            "source_name": clean(r.get("source_name")),
            "source_url": clean(r.get("source_url")),
            "is_valuable": bool(r.get("is_valuable")),
            "score": clean(r.get("score")),
            "priority": int(r.get("priority") or 0),
        })

    # ---- 同坐标散开：同坐标组内"不同名"项目微偏移（前端 ≤3 条不折叠 → 散开后各自显示不重叠；
    #      同名/相似项目不散开，保持同坐标 → 前端同名折叠防重） ----
    from collections import defaultdict
    coord_groups = defaultdict(list)
    for i, p in enumerate(map_points):
        key = (round(p["_lng"] * 1e4), round(p["_lat"] * 1e4))
        coord_groups[key].append(i)
    for idxs in coord_groups.values():
        if len(idxs) < 2:
            continue
        base_name = map_points[idxs[0]]["name"]
        for j, i in enumerate(idxs[1:], 1):
            nm = map_points[i]["name"]
            # 与组内第一个同名/相似（前缀包含）→ 不散开（前端同名折叠）
            if nm and base_name and (nm == base_name or nm[:8] == base_name[:8]
                                     or base_name[:8] in nm or nm[:8] in base_name):
                continue
            map_points[i]["value"][0] = round(map_points[i]["_lng"] + 0.002 + (j % 3) * 0.002, 6)
            map_points[i]["value"][1] = round(map_points[i]["_lat"] + (j % 2) * 0.002, 6)
    for p in map_points:
        p.pop("_lng", None)
        p.pop("_lat", None)

    # ---- project_list：完整内容（仅红/黄预警，与 map_points 同口径；无预警不入库） ----
    project_list = []
    for r in rows:
        if "红色" not in str(r.get("warning_level", "")) \
                and "黄色" not in str(r.get("warning_level", "")):
            continue
        project_list.append({
            "name": clean(r.get("project_name")),
            "district": clean(r.get("district")),
            "type": clean(r.get("project_type")),
            "stage": stage_of(r.get("status")),              # 5 类阶段（周期过滤/分组统计）
            "stage_detail": clean(r.get("status")),          # AI 完整阶段描述（详情兜底）
            "priority": int(r.get("priority") or 0),
            "warning": clean(r.get("warning_level")),
            "investment": clean(r.get("investment")),
            "scale": clean(r.get("scale")),
            "date": norm_date(r.get("publish_date")),
            "url": clean(r.get("source_url")),
            "location": clean(r.get("location")),
            "ai_summary": clean(r.get("ai_summary")),
            "processed": 1 if (r.get("processed_status") or 0) else 0,
            "location": clean(r.get("location")),
            "developer": clean(r.get("developer")),
            "ai_reason": clean(r.get("ai_reason")),
            "content": clean(r.get("content")),
            "telecom_needs": clean(r.get("telecom_needs")),
            "is_valuable": bool(r.get("is_valuable")),
        })

    return {
        "meta": {
            "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "total_projects": total,
            "source": "MySQL 数据库导出（export_dashboard_db.py）",
        },
        "summary": {
            "total": total,
            "red_warning": red,
            "yellow_warning": yellow,
            "valuable": valuable,
            "district_count": len(districts),
        },
        "warning_pie": warning_pie,
        "type_pie": type_pie,
        "stage_pie": stage_pie,
        "district_ranking": district_ranking,
        "timeline": timeline,
        "red_timeline": red_timeline,
        "yellow_timeline": yellow_timeline,
        "map_points": map_points,
        "project_list": project_list,
    }


DEFAULT_DASHBOARD_OUTPUT = str(
    Path(__file__).resolve().parent.parent / "frontend" / "data" / "dashboard_data.json")
DEFAULT_REPORT_OUTPUT = str(
    Path(__file__).resolve().parent.parent / "frontend" / "data" / "report_data.json")


def export_dashboard(output: str = None) -> str:
    """MySQL → 大屏 dashboard_data.json（全量），返回输出路径（供 scheduler 自动化调用）"""
    rows = fetch_projects()
    print(f"[DB] 读取 {len(rows)} 条记录")
    data = build_dashboard(rows)
    out = Path(output or DEFAULT_DASHBOARD_OUTPUT)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"[OK] 已导出 {out}（map_points {len(data['map_points'])} / project_list {len(data['project_list'])}）")
    return str(out)


def export_report(output: str = None) -> str:
    """MySQL → 报告 report_data.json（全量，供 report.html 全年月度趋势 + 较上月对比），返回输出路径

    说明：月度趋势折线图、统计卡片"较上月"对比都需要全年数据，
    因此与 dashboard 一样导出全部记录（不做月份筛选）。
    """
    rows = fetch_projects()
    data = build_dashboard(rows)
    out = Path(output or DEFAULT_REPORT_OUTPUT)
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=1)
    print(f"[OK] 已导出 {out}（map_points {len(data['map_points'])} / project_list {len(data['project_list'])}）")
    return str(out)


def fetch_weekly_stats() -> dict:
    """查询本周项目统计（publish_date >= 本周一，自然周口径），供周报 Excel / 邮件正文使用"""
    import datetime as _dt

    rows = fetch_projects()
    today = _dt.date.today()
    week_ago = today - _dt.timedelta(days=today.weekday())  # 本周一（weekday: 0=周一）

    week_rows = []
    for r in rows:
        d = r.get("publish_date") or ""
        try:
            pd = _dt.date.fromisoformat(str(d)[:10])
        except ValueError:
            continue
        if pd >= week_ago:
            week_rows.append(r)

    dist_map = {}
    red = yellow = 0
    for r in week_rows:
        d = r.get("district") or "未知"
        wl = r.get("warning_level") or ""
        e = dist_map.setdefault(d, {"total": 0, "red": 0, "yellow": 0})
        e["total"] += 1
        if "红" in wl:
            e["red"] += 1
            red += 1
        elif "黄" in wl:
            e["yellow"] += 1
            yellow += 1

    by_district = [{"district": d, **e} for d, e in
                   sorted(dist_map.items(), key=lambda kv: -kv[1]["total"])]
    return {
        "total": len(week_rows),
        "red": red,
        "yellow": yellow,
        "district_count": len(dist_map),
        "week_ago": week_ago.isoformat(),
        "today": today.isoformat(),
        "by_district": by_district,
        "rows": week_rows,
    }


def export_weekly_excel(output: str = None) -> str:
    """生成周报 Excel（概览 + 本周项目明细 + 区县柱状图原生图表），供邮件附件

    布局（给领导一目了然）：
      A1 大标题「工程建设信息预警周报」
      A2 报告周期
      A3 本周统计（项目总数/红/黄/覆盖区县）
      第 5 行起：本周项目明细（自动筛选）
      M 列：区县汇总数据；M20 起：原生柱状图（区县横坐标）
    """
    import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

    stats = fetch_weekly_stats()
    week_rows = stats["rows"]
    by_district = stats["by_district"]

    wb = Workbook()
    ws = wb.active
    ws.title = "本周项目"

    title_font = Font(name="微软雅黑", size=18, bold=True, color="1F4E79")
    sub_font = Font(name="微软雅黑", size=11, color="404040")
    stat_font = Font(name="微软雅黑", size=12, bold=True, color="C00000")

    # ---- 概览区 ----
    ws.merge_cells("A1:K1")
    ws["A1"] = "工程建设信息预警周报"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:K2")
    ws["A2"] = f"报告周期：{stats['week_ago']} 至 {stats['today']}"
    ws["A2"].font = sub_font

    ws.merge_cells("A3:K3")
    ws["A3"] = (
        f"本周项目 {stats['total']} 个 ｜ 红色预警 {stats['red']} 个 ｜ "
        f"黄色预警 {stats['yellow']} 个 ｜ 覆盖区县 {stats['district_count']} 个"
    )
    ws["A3"].font = stat_font
    ws.row_dimensions[3].height = 24

    # ---- 本周项目明细（第 5 行起） ----
    headers = ["项目名称", "区县", "类型", "阶段", "预警级别", "发布日期",
               "地点", "投资额", "规模", "来源", "原文链接"]
    HEADER_ROW = 5
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=HEADER_ROW, column=ci, value=h)
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[HEADER_ROW].height = 22

    for ri, r in enumerate(week_rows):
        d = r.get("publish_date") or ""
        try:
            dstr = _dt.date.fromisoformat(str(d)[:10]).isoformat()
        except Exception:
            dstr = ""
        vals = [r.get("project_name") or "", r.get("district") or "",
                r.get("project_type") or "", r.get("status") or "",
                r.get("warning_level") or "", dstr,
                r.get("location") or "", r.get("investment") or "",
                r.get("scale") or "", r.get("source_name") or "",
                r.get("source_url") or ""]
        for ci, v in enumerate(vals, 1):
            ws.cell(row=HEADER_ROW + 1 + ri, column=ci, value=v)

    last_data = HEADER_ROW + len(week_rows)
    ws.auto_filter.ref = f"A{HEADER_ROW}:K{last_data}"
    ws.freeze_panes = f"A{HEADER_ROW + 1}"
    for col, w in zip(range(1, 12), [42, 10, 12, 14, 10, 12, 38, 16, 16, 14, 55]):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ---- Sheet2：区县汇总（独立 sheet，供 Sheet1 柱状图引用） ----
    ws2 = wb.create_sheet("区县汇总")
    ws2.append(["区县", "项目数", "红色预警", "黄色预警"])
    for e in by_district:
        ws2.append([e["district"], e["total"], e["red"], e["yellow"]])
    for cell in ws2[1]:
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")
    ws2.column_dimensions["A"].width = 14
    for c in "BCD":
        ws2.column_dimensions[c].width = 12

    out = Path(output or (Path(__file__).resolve().parent.parent / "data" / "weekly_excel" / (
        f"工程建设信息预警周报_{stats['today']}.xlsx")))
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] 周报 Excel 已生成: {out}（本周 {stats['total']} 条）")
    return str(out)


def export_all_excel(output: str = None) -> str:
    """数据库全部项目 → Excel（项目明细 + 区县汇总 + 类型汇总），带自动筛选"""
    import datetime as _dt
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    rows = fetch_projects()
    today = _dt.date.today()

    wb = Workbook()
    ws = wb.active
    ws.title = "项目明细"

    headers = ["项目名称", "区县", "类型", "阶段", "预警级别", "基站需求", "基站类型",
               "覆盖范围", "发布日期", "投标截止", "预计开工", "预计竣工", "地点",
               "投资额", "规模", "建设单位", "联系人", "联系电话", "通信需求",
               "优先级", "评分", "是否高价值", "AI摘要", "AI推理", "公告原文",
               "处理状态", "备注", "来源", "原文链接", "创建时间", "更新时间"]
    ws.append(headers)
    ws.freeze_panes = "A2"
    for r in rows:
        needs = r.get("telecom_needs") or ""
        d = r.get("publish_date") or ""
        try:
            dstr = _dt.date.fromisoformat(str(d)[:10]).isoformat()
        except Exception:
            dstr = ""
        ws.append([
            r.get("project_name") or "", r.get("district") or "",
            r.get("project_type") or "", r.get("status") or "",
            r.get("warning_level") or "", r.get("need_base_station") or "",
            r.get("base_station_type") or "", r.get("coverage_area") or "",
            dstr, r.get("deadline") or "", r.get("start_date") or "",
            r.get("end_date") or "", r.get("location") or "",
            r.get("investment") or "", r.get("scale") or "",
            r.get("developer") or "", r.get("contact_person") or "",
            r.get("contact_phone") or "", needs,
            r.get("priority") or "", r.get("score") or "",
            "是" if r.get("is_valuable") else "否",
            r.get("ai_summary") or "", r.get("ai_reason") or "",
            r.get("content") or "",
            "已处理" if r.get("processed_status") else "未处理",
            r.get("notes") or "", r.get("source_name") or "",
            r.get("source_url") or "",
            str(r.get("created_at"))[:19] if r.get("created_at") else "",
            str(r.get("updated_at"))[:19] if r.get("updated_at") else "",
        ])
    ncols = len(headers)
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{len(rows) + 1}"
    for col, w in zip(range(1, ncols + 1), [45, 10, 12, 14, 10, 10, 12, 30, 12, 12, 12, 12, 38, 16, 16, 16, 10, 14, 30, 8, 8, 10, 50, 60, 50, 10, 20, 14, 55, 20, 20]):
        ws.column_dimensions[get_column_letter(col)].width = w
    for cell in ws[1]:
        cell.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E79")

    # 区县汇总 sheet
    ws2 = wb.create_sheet("区县汇总")
    dist_map = {}
    for r in rows:
        d = r.get("district") or "未知"
        wl = r.get("warning_level") or ""
        e = dist_map.setdefault(d, {"total": 0, "red": 0, "yellow": 0})
        e["total"] += 1
        if "红" in wl:
            e["red"] += 1
        elif "黄" in wl:
            e["yellow"] += 1
    ws2.append(["区县", "项目数", "红色预警", "黄色预警"])
    for d, e in sorted(dist_map.items(), key=lambda kv: -kv[1]["total"]):
        ws2.append([d, e["total"], e["red"], e["yellow"]])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    ws2.column_dimensions["A"].width = 14
    for c in "BCD":
        ws2.column_dimensions[c].width = 12

    # 类型汇总 sheet
    ws3 = wb.create_sheet("类型汇总")
    type_map = {}
    for r in rows:
        t = r.get("project_type") or "未知"
        type_map[t] = type_map.get(t, 0) + 1
    ws3.append(["项目类型", "项目数"])
    for t, c in sorted(type_map.items(), key=lambda kv: -kv[1]):
        ws3.append([t, c])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 10

    out = Path(output or (Path(__file__).resolve().parent.parent / "data" / "weekly_excel" / (
        f"数据库全量项目_{today.isoformat()}.xlsx")))
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"[OK] 全量 Excel 已生成: {out}（{len(rows)} 条）")
    return str(out)


def main():
    parser = argparse.ArgumentParser(description="MySQL → 大屏 dashboard_data.json")
    parser.add_argument("-o", "--output", default=DEFAULT_DASHBOARD_OUTPUT)
    parser.add_argument("--report", action="store_true",
                        help="导出报告数据（仅本月+本周）到 frontend/data/report_data.json")
    parser.add_argument("--weekly-excel", action="store_true",
                        help="生成周报 Excel（本周项目 + 区县汇总）")
    parser.add_argument("--all-excel", action="store_true",
                        help="生成数据库全量项目 Excel（明细 + 区县/类型汇总）")
    args = parser.parse_args()

    if args.all_excel:
        export_all_excel(args.output if args.output != DEFAULT_DASHBOARD_OUTPUT else None)
    elif args.weekly_excel:
        export_weekly_excel(args.output if args.output != DEFAULT_DASHBOARD_OUTPUT else None)
    elif args.report:
        export_report(args.output if args.output != DEFAULT_DASHBOARD_OUTPUT else None)
    else:
        export_dashboard(args.output)


if __name__ == "__main__":
    main()
