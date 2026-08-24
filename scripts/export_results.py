"""
AI 分析结果 → Excel 格式化导出
-------------------------------
将 ai_pipeline.py 输出的 JSON（unified_intelligence.json）导出为
带颜色标记、冻结表头的审核用 Excel 文件。

用法：
  python scripts/export_results.py data/unified_intelligence.json
  python scripts/export_results.py data/unified_intelligence.json -o data/results.xlsx
"""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def export_ai_results(results: list[dict], output_path: str) -> None:
    """将 AI 分析结果导出为格式化 Excel 文件。"""
    if not results:
        print("[EMPTY] 无数据可导出")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "AI情报分析结果"

    # ---- 样式定义 ----
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ---- 表头（与 review_data.py 一致，评分→预警等级，质量→项目阶段） ----
    headers = [
        "序号", "预警等级", "项目阶段", "标题", "区县", "规模", "投资额",
        "项目性质", "发布日期", "URL", "内容摘要(前200字)", "审核结果",
    ]
    col_widths = [6, 10, 12, 45, 10, 12, 12, 8, 12, 35, 40, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    # ---- 数据行 ----
    for i, item in enumerate(results, 1):
        row = i + 1

        warning = item.get("warning_level", "")
        project_stage = item.get("project_stage", "")
        title = item.get("project_name", "")
        district = item.get("district", "")
        scale = item.get("scale", "")
        investment = item.get("investment", "")
        project_type = item.get("project_type", "")
        pub_date = item.get("_publish_date", "")
        url = item.get("_source_url", "")
        ai_summary = item.get("ai_summary", "")

        # 内容摘要：取 ai_summary 前200字
        summary = ai_summary.replace("\n", " ")[:200] if ai_summary else ""

        # 行颜色（按预警等级）
        if "红色" in str(warning):
            row_fill = red_fill
        elif "黄色" in str(warning):
            row_fill = yellow_fill
        else:
            row_fill = green_fill

        row_data = [
            i, warning, project_stage, title, district, scale, investment,
            project_type, pub_date, url, summary, "",
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if col_idx != 12:  # 审核结果列不填色
                cell.fill = row_fill

        # URL 列超链接
        url_cell = ws.cell(row=row, column=10)
        if url and url.startswith("http"):
            url_cell.font = link_font
            url_cell.hyperlink = url

        ws.row_dimensions[row].height = 28

    # ---- 冻结首行 + 自动筛选 ----
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L{len(results) + 1}"

    # ---- 使用说明 sheet ----
    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions["A"].width = 80
    instructions = [
        "AI 情报分析结果说明",
        "",
        "颜色说明:",
        "  红色  = 红色预警（优先级≥4，需立即跟进）",
        "  黄色  = 黄色预警（优先级3，建议跟进）",
        "  绿色  = 无预警 / 低优先级",
        "",
        "列说明:",
        "  预警等级: AI 综合评估的紧急程度（红色>黄色>无预警）",
        "  项目阶段: 规划立项/招标阶段/施工阶段/已竣工完工/待核实",
        "  标题: 核实后的标准项目名称",
        "  区县/规模/投资额/项目性质: AI 从公告原文提取",
        "  内容摘要: AI 对该项目的一句话情报总结",
        "",
        "审核方法:",
        "  1. 按预警等级排序（红色优先）",
        "  2. 点击 URL 查看政府公告原文",
        "  3. 确认项目阶段和建设方信息",
        "  4. 在「审核结果」列标记，安排销售跟进",
    ]
    for row_idx, text in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(name="微软雅黑", size=14, bold=True)
        elif text.startswith("  ") and not text.startswith("    "):
            cell.font = Font(name="微软雅黑", size=10, bold=True)
        else:
            cell.font = Font(name="微软雅黑", size=10)

    # ---- 保存 ----
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"\n[OK] AI 结果 Excel 导出完成!")
    print(f"   文件: {output_path}")
    print(f"   共 {len(results)} 条数据")
    print(f"   包含「{ws.title}」和「使用说明」两个 sheet")


def main():
    parser = argparse.ArgumentParser(
        description="AI 分析结果 JSON → 格式化 Excel 导出",
    )
    parser.add_argument(
        "input_json",
        help="AI 分析结果 JSON 文件（unified_intelligence.json）",
    )
    parser.add_argument(
        "--output", "-o", type=str, default=None,
        help="输出 Excel 路径（默认：输入文件名.xlsx）",
    )
    args = parser.parse_args()

    with open(args.input_json, "r", encoding="utf-8") as f:
        results = json.load(f)

    output_path = args.output
    if not output_path:
        p = Path(args.input_json)
        output_path = str(p.parent / f"{p.stem}.xlsx")

    export_ai_results(results, output_path)


if __name__ == "__main__":
    main()
