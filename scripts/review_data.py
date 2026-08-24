"""
数据审核脚本
-----------
爬虫测试数据 → 人工审核 → 写入数据库（触发 AI 分析）

工作流:
  1. python scripts/test_spider.py yantai_bidding   # 爬取测试
  2. python scripts/review_data.py data/spider_test/xxx.json  # 逐条审核
  3. python scripts/review_data.py data/spider_test/xxx.json --approve  # 批量通过
  4. python scripts/review_data.py data/spider_test/xxx.json --export-excel  # 导出Excel

设计原则: 用户可以控制哪些数据进入 AI 处理管线。
"""

import sys
import json
import datetime
import csv
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Flask/SQLAlchemy 懒加载 — export_excel 不需要数据库
_flask_imports_done = False


def _ensure_db_imports():
    """延迟导入数据库模块（仅写入数据库时需要）"""
    global _flask_imports_done
    if _flask_imports_done:
        return
    import flask  # noqa: F401
    from config.settings import DATABASE_URL, FLASK_SECRET_KEY  # noqa: F401
    from database.models import db, RawProject  # noqa: F401
    from database.db_manager import add_raw_project, count_unprocessed  # noqa: F401
    _flask_imports_done = True


def load_data(filepath: str) -> list:
    """加载爬虫测试 JSON 文件"""
    path = Path(filepath)
    if not path.exists():
        print(f"[ERROR] 文件不存在: {path}")
        sys.exit(1)

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    print(f"[LOAD] 加载 {len(data)} 条数据: {path.name}")
    return data


def print_item(i: int, item: dict):
    """打印单条数据的关键信息"""
    score = item.get("relevance_score", "?")
    district = item.get("district_extracted", "")
    scale = item.get("scale_extracted", "")
    investment = item.get("investment_extracted", "")
    nature = item.get("project_nature", "")
    date = item.get("publish_date", "")

    # 质量图标
    if isinstance(score, int):
        icon = "[高]" if score >= 5 else "[中]" if score >= 3 else "[低]" if score >= 1 else "[弃]"
    else:
        icon = "[?]"

    print(f"\n{'─'*70}")
    print(f"  [{i}] {icon} 评分: {score}  |  {item.get('title', 'N/A')[:70]}")
    print(f"{'─'*70}")

    extras = []
    if district:
        extras.append(f"区:{district}")
    if scale:
        extras.append(f"规模:{scale}")
    if investment:
        extras.append(f"投资:{investment}")
    if nature:
        extras.append(f"性质:{nature}")
    if date:
        extras.append(f"日期:{date}")

    if extras:
        print(f"  {' | '.join(extras)}")

    url = item.get("source_url", "")
    if url:
        print(f"  URL: {url[:100]}")

    # 内容摘要
    content = item.get("content", "")
    if content:
        print(f"  内容: {(content.replace(chr(10), ' '))[:150]}...")


def interactive_review(filepath: str):
    """交互式逐条审核"""
    data = load_data(filepath)
    approved = []
    rejected = []
    skipped = []

    print(f"\n[审核] 开始逐条审核 (y=通过 / n=拒绝 / s=跳过 / q=退出)")
    print(f"{'='*70}")

    for i, item in enumerate(data, 1):
        print_item(i, item)

        while True:
            choice = input(f"\n  > [{i}/{len(data)}] y/n/s/q? ").strip().lower()
            if choice == "y":
                approved.append(item)
                print(f"    [Y] 已通过")
                break
            elif choice == "n":
                rejected.append(item)
                print(f"    [N] 已拒绝")
                break
            elif choice == "s":
                skipped.append(item)
                print(f"    [S] 已跳过")
                break
            elif choice == "q":
                # 剩余的全部跳过
                remaining = len(data) - i
                if remaining > 0:
                    skipped.extend(data[i:])
                    print(f"    [S] 剩余 {remaining} 条跳过")
                break
            else:
                print(f"    请输入 y/n/s/q")

        if choice == "q":
            break

    # ---- 汇总 ----
    print(f"\n{'='*70}")
    print(f"审核完成:")
    print(f"  [Y] 通过: {len(approved)} 条")
    print(f"  [N] 拒绝: {len(rejected)} 条")
    print(f"  [S] 跳过: {len(skipped)} 条")
    print(f"{'='*70}")

    if approved:
        print(f"\n准备将 {len(approved)} 条通过的数据写入数据库...")
        confirm = input("确认写入? (y/n): ").strip().lower()
        if confirm == "y":
            write_to_db(approved)


def bulk_approve(filepath: str):
    """批量通过（--approve 模式）"""
    data = load_data(filepath)

    # 统计
    high = sum(1 for d in data if d.get("relevance_score", 0) >= 5)
    mid = sum(1 for d in data if 3 <= d.get("relevance_score", 0) < 5)
    low = sum(1 for d in data if 0 < d.get("relevance_score", 0) < 3)

    print(f"\n[统计] 数据质量分布:")
    print(f"  [高] 高质量 (>=5): {high} 条")
    print(f"  [中] 中等 (3-4):   {mid} 条")
    print(f"  [低] 低质量 (1-2): {low} 条")

    # 可选: 只写入高质量+中等
    print(f"\n写入策略:")
    print(f"  1) 全部写入 ({len(data)} 条)")
    print(f"  2) 只写入 >= 3 分 ({high + mid} 条)")
    print(f"  3) 只写入 >= 5 分 ({high} 条)")
    print(f"  4) 取消")

    choice = input("选择 (1/2/3/4): ").strip()

    if choice == "1":
        to_write = data
    elif choice == "2":
        to_write = [d for d in data if d.get("relevance_score", 0) >= 3]
    elif choice == "3":
        to_write = [d for d in data if d.get("relevance_score", 0) >= 5]
    else:
        print("已取消")
        return

    if to_write:
        write_to_db(to_write)


def export_excel(filepath: str):
    """导出审核用 Excel 文件（含评分、区县、规模等列，方便人工审核）"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("[ERROR] 需要安装 openpyxl: pip install openpyxl")
        return

    data = load_data(filepath)
    if not data:
        print("[EMPTY] 无数据可导出")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "爬虫数据审核"

    # ---- 样式定义 ----
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    high_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")    # 绿色-高质量
    mid_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")     # 黄色-中等
    low_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # 红色-低质量
    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # ---- 表头 ----
    headers = ["序号", "来源", "评分", "质量", "标题", "区县", "规模", "投资额",
               "项目性质", "发布日期", "URL", "内容摘要(前200字)", "审核结果"]
    col_widths = [6, 12, 6, 8, 45, 10, 12, 12, 8, 12, 35, 40, 10]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    # ---- 数据行 ----
    for i, item in enumerate(data, 1):
        row = i + 1
        score = item.get("relevance_score", 0)
        district = item.get("district_extracted", "")
        scale = item.get("scale_extracted", "")
        investment = item.get("investment_extracted", "")
        nature = item.get("project_nature", "")
        date = item.get("publish_date", "")
        title = item.get("title", "")
        url = item.get("source_url", "")
        content = item.get("content", "")

        # 质量标签
        if isinstance(score, int):
            if score >= 5:
                quality = "高质量"
                row_fill = high_fill
            elif score >= 3:
                quality = "中等"
                row_fill = mid_fill
            elif score >= 1:
                quality = "低质量"
                row_fill = low_fill
            else:
                quality = "应丢弃"
                row_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        else:
            quality = "未知"
            row_fill = None

        # 内容摘要
        summary = content.replace("\n", " ")[:200] if content else ""

        row_data = [i, item.get("source", ""), score, quality, title, district,
                    scale, investment, nature, date, url, summary, ""]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            if row_fill and col_idx != 12:  # 审核结果列不填色
                cell.fill = row_fill

        # URL 列超链接样式
        url_cell = ws.cell(row=row, column=10)
        if url and url.startswith("http"):
            url_cell.font = link_font
            url_cell.hyperlink = url

        ws.row_dimensions[row].height = 28

    # ---- 冻结首行 ----
    ws.freeze_panes = "A2"

    # ---- 自动筛选 ----
    ws.auto_filter.ref = f"A1:L{len(data)+1}"

    # ---- 添加说明 sheet ----
    ws2 = wb.create_sheet("使用说明")
    ws2.column_dimensions["A"].width = 80
    instructions = [
        "数据审核 Excel 使用说明",
        "",
        "列说明:",
        "  评分: 相关性评分（越高越可能是大型建设工程）",
        "  质量: [高]高质量(>=5) [中]中等(3-4) [低]低质量(1-2) [弃]应丢弃(<=0)",
        "  区县/规模/投资额/性质: 爬虫从正文自动提取",
        "",
        "审核方法:",
        "  1. 逐行查看标题、评分、规模",
        "  2. 点击 URL 查看原文",
        "  3. 在「审核结果」列填写: 通过 / 拒绝 / 待确认",
        "  4. 审核完成后，通过的数据可写入数据库",
        "",
        "注意事项:",
        "  - 评分是辅助工具，不绝对准确 — 建议人工确认",
        "  - 低分但有规模的可能是漏网之鱼，高分但实际是小改造的需排除",
        "  - URL 列可 Ctrl+点击 跳转到原文",
    ]
    for row, text in enumerate(instructions, 1):
        cell = ws2.cell(row=row, column=1, value=text)
        if row == 1:
            cell.font = Font(name="微软雅黑", size=14, bold=True)
        elif text.startswith("  ") and not text.startswith("    "):
            cell.font = Font(name="微软雅黑", size=10, bold=True)
        else:
            cell.font = Font(name="微软雅黑", size=10)

    # ---- 保存 ----
    path = Path(filepath)
    excel_path = path.parent / f"{path.stem}_审核表.xlsx"

    wb.save(excel_path)
    print(f"\n[OK] Excel 导出完成!")
    print(f"   文件: {excel_path}")
    print(f"   共 {len(data)} 条数据")
    print(f"   包含「{ws.title}」和「使用说明」两个 sheet")
    print(f"\n提示: 打开后在「审核结果」列标记，完成后用 --approve 写入数据库")


def write_to_db(items: list):
    """将审核通过的数据写入 raw_projects 表"""
    _ensure_db_imports()
    from flask import Flask
    from config.settings import DATABASE_URL, FLASK_SECRET_KEY
    from database.models import db, RawProject
    from database.db_manager import add_raw_project, count_unprocessed

    # 初始化 Flask + 数据库
    app = Flask(__name__)
    app.config["SECRET_KEY"] = FLASK_SECRET_KEY
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    db.init_app(app)

    with app.app_context():
        db.create_all()

        written = 0
        skipped_dup = 0

        for item in items:
            title = item.get("title", "")
            source_url = item.get("source_url", "")

            # 去重检查
            existing = RawProject.query.filter_by(
                title=title,
                source_name="烟台市公共资源交易网",
            ).first()

            if existing:
                skipped_dup += 1
                continue

            # 将评分信息附加到 content 中保存
            content = item.get("content", "")
            score = item.get("relevance_score", 0)
            district = item.get("district_extracted", "")
            scale = item.get("scale_extracted", "")
            nature = item.get("project_nature", "")

            # 构建增强的 content（包含爬虫预提取信息）
            meta_parts = []
            if district:
                meta_parts.append(f"区县: {district}")
            if scale:
                meta_parts.append(f"规模: {scale}")
            if nature:
                meta_parts.append(f"性质: {nature}")
            meta_parts.append(f"相关性评分: {score}")

            enhanced_content = f"[爬虫预提取]\n" + "\n".join(meta_parts)
            if content:
                enhanced_content += f"\n\n[正文]\n{content}"

            add_raw_project(
                title=title,
                source_name="烟台市公共资源交易网",
                content=enhanced_content,
                source_url=source_url,
                publish_date=item.get("publish_date", ""),
            )
            written += 1

        print(f"\n[OK] 写入完成:")
        print(f"  新增: {written} 条")
        print(f"  重复跳过: {skipped_dup} 条")
        print(f"  待处理总数: {count_unprocessed()} 条")

    if written > 0:
        print(f"\n提示: 下一步: 运行 AI 处理管线")
        print(f"   python -c \"from crawler.scheduler import main; main()\"")
        print(f"   或者单独运行 AI 处理:")
        print(f"   (需要先确认豆包 API 配置正确)")


def main():
    if len(sys.argv) < 2:
        print("数据审核脚本")
        print()
        print("用法:")
        print("  python scripts/review_data.py <JSON文件>               # 交互式逐条审核")
        print("  python scripts/review_data.py <JSON文件> --approve     # 批量审核写入数据库")
        print("  python scripts/review_data.py <JSON文件> --export-excel # 导出Excel审核表")
        print()
        print("示例:")
        print("  python scripts/review_data.py data/spider_test/yantai_bidding_xxx.json")
        print("  python scripts/review_data.py data/spider_test/yantai_bidding_xxx.json --approve")
        print("  python scripts/review_data.py data/spider_test/yantai_bidding_xxx.json --export-excel")
        sys.exit(0)

    filepath = sys.argv[1]
    bulk = "--approve" in sys.argv
    export = "--export-excel" in sys.argv

    if export:
        export_excel(filepath)
    elif bulk:
        bulk_approve(filepath)
    else:
        interactive_review(filepath)


if __name__ == "__main__":
    main()
