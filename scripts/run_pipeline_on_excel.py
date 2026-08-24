"""
读取"各个区县数据.xlsx" → AI 管线批量分析 → workbuddy.json → 导入 MySQL

用法:
  python scripts/run_pipeline_on_excel.py
  python scripts/run_pipeline_on_excel.py --limit 10   # 测试模式：只跑前N条
"""
import sys, os, json, time, argparse
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from loguru import logger
from processor.doubao_client import doubao
from processor.ai_pipeline import UNIFIED_SYSTEM_PROMPT, build_user_message, _clean_unified_result
from config.settings import YANTAI_DISTRICTS

EXCEL_PATH = r"D:\网络部工作\新建文件夹\各个区县数据.xlsx"
OUTPUT_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                           "data", "results", "workbuddy.json")

# --- 1. 读取 Excel ---
wb = openpyxl.load_workbook(EXCEL_PATH)
ws = wb[wb.sheetnames[0]]
print(f"Excel: {ws.max_row} 行（含表头）, {ws.max_column} 列")

records = []
for row_idx in range(2, ws.max_row + 1):  # 跳过表头
    rec = {
        "_excel_row": row_idx,
        "_index": ws.cell(row=row_idx, column=1).value,
        "_spider_score": ws.cell(row=row_idx, column=2).value,
        "_spider_quality": ws.cell(row=row_idx, column=3).value,
        "title": str(ws.cell(row=row_idx, column=4).value or ""),
        "district": str(ws.cell(row=row_idx, column=5).value or ""),
        "scale": str(ws.cell(row=row_idx, column=6).value or "") if ws.cell(row=row_idx, column=6).value else "",
        "investment": str(ws.cell(row=row_idx, column=7).value or "") if ws.cell(row=row_idx, column=7).value else "",
        "project_nature": str(ws.cell(row=row_idx, column=8).value or "") if ws.cell(row=row_idx, column=8).value else "",
        "publish_date": str(ws.cell(row=row_idx, column=9).value or "")[:10],
        "source_url": str(ws.cell(row=row_idx, column=10).value or ""),
        "content": str(ws.cell(row=row_idx, column=11).value or ""),
        "review_result": str(ws.cell(row=row_idx, column=12).value or "") if ws.cell(row=row_idx, column=12).value else "",
        "source_name": "yantai_districts_excel",
    }
    # 跳过空行
    if not rec["title"] or rec["title"] == "None":
        continue
    records.append(rec)

print(f"有效记录: {len(records)} 条")

# --- 参数解析 ---
parser = argparse.ArgumentParser()
parser.add_argument("--limit", type=int, default=0, help="只跑前N条（测试用）")
parser.add_argument("--index", type=int, default=0, help="只跑指定序号的单条（测试用）")
args = parser.parse_args()

if args.limit and args.limit > 0:
    records = records[:args.limit]
    print(f"⚡ 测试模式：仅处理前 {len(records)} 条")

if args.index and args.index > 0:
    records = [r for r in records if r.get("_index") == args.index]
    print(f"⚡ 单条模式：仅处理 _index={args.index}（剩余 {len(records)} 条）")

# --- 2. 检查已有结果（断点续跑） ---
existing_results = []
existing_indices = set()
if os.path.exists(OUTPUT_PATH):
    try:
        with open(OUTPUT_PATH, "r", encoding="utf-8") as f:
            existing_results = json.load(f)
        existing_indices = {r.get("_index") for r in existing_results if r.get("_index")}
        if existing_indices:
            print(f"已有结果: {len(existing_indices)} 条，将跳过")
    except Exception:
        print("已有 workbuddy.json 损坏，重新开始")

# --- 3. 批量分析 ---
doubao.reset_cumulative_usage()
t_start = time.time()
total = len(records)
success = 0
fail = 0
skip = 0
red_count = 0
yellow_count = 0
results = existing_results.copy()

for i, rec in enumerate(records):
    idx = rec.get("_index")

    # 断点续跑：跳过已有
    if idx in existing_indices:
        skip += 1
        continue

    title = rec["title"]
    content = rec["content"]
    url = rec["source_url"]
    date_str = rec.get("publish_date", "")
    district = rec.get("district", "")

    preview = title[:60].replace("\n", " ")
    elapsed_total = time.time() - t_start
    eta = (elapsed_total / max(1, i - skip)) * (total - i) if i > skip else 0
    print(f"[{i+1}/{total}] {preview}...", end=" ", flush=True)

    user_msg = build_user_message(
        title, content, date_str, url, rec["source_name"],
    )
    # 把爬虫已提取的区县作为提示喂给 AI（仅供参考，仍需联网核实）
    if district:
        user_msg += f"\n\n## 爬虫已提取区县（仅供参考，请联网核实后填写 district）\n{district}"

    # 注入烟台标准区县白名单，约束 AI 的 district 取值（防止编造白名单外区县）
    district_hint = "、".join(YANTAI_DISTRICTS)
    system_prompt = (
        UNIFIED_SYSTEM_PROMPT
        + f"\n\n## 烟台标准区县白名单（district 必须且只能是以下之一，无法确认填'待核实'）\n"
        + district_hint
    )

    t0 = time.time()
    raw = doubao.chat(
        system_prompt=system_prompt,
        user_message=user_msg,
        temperature=0.0,
        max_tokens=4096,
        enable_web_search=True,
    )
    elapsed = time.time() - t0

    if raw is None:
        print(f"❌ API返回空")
        fail += 1
        continue

    # 解析 JSON
    try:
        ai_result = json.loads(raw)
    except json.JSONDecodeError:
        import re
        m = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', raw)
        if m:
            try:
                ai_result = json.loads(m.group(1))
            except json.JSONDecodeError:
                print(f"❌ JSON解析失败: {raw[:80]}")
                fail += 1
                continue
        else:
            print(f"❌ JSON解析失败: {raw[:80]}")
            fail += 1
            continue

    # 清洗结果
    ai_result = _clean_unified_result(ai_result)
    # 统一 need_base_station 为 有/无（匹配现有 workbuddy.json 格式）
    nb = ai_result.get("need_base_station")
    ai_result["need_base_station"] = "无" if nb == "无" else "有"
    ai_result["_source_url"] = url
    ai_result["_source_name"] = rec["source_name"]
    ai_result["_publish_date"] = date_str
    ai_result["_title"] = title
    ai_result["_index"] = idx
    ai_result["_spider_score"] = rec.get("_spider_score")
    ai_result["_spider_quality"] = rec.get("_spider_quality")

    wl = ai_result.get("warning_level", "?")
    if wl == "红色预警":
        red_count += 1
    elif wl == "黄色预警":
        yellow_count += 1

    is_skip = ai_result.get("skip", False)
    status_icon = "⏭️" if is_skip else "✅"
    print(f"{status_icon} {elapsed:.1f}s | {wl} | {ai_result.get('project_name','?')[:30]}")

    results.append(ai_result)
    success += 1

    # 每 10 条保存一次（防止中断丢失）
    if success % 10 == 0:
        with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        usage = doubao.cumulative_usage
        print(f"  💾 已保存 | 成功:{success} 失败:{fail} 跳过:{skip} | "
              f"累计 {usage['total_tokens']} tokens | ETA {eta/60:.0f}min")

# --- 4. 最终保存 ---
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

total_elapsed = time.time() - t_start
usage = doubao.cumulative_usage
valid_results = [r for r in results if not r.get("skip")]
skip_results = [r for r in results if r.get("skip")]

print()
print("=" * 60)
print(f"🏁 完成！")
print(f"  总耗时: {total_elapsed/60:.1f} 分钟")
print(f"  成功: {success} | 失败: {fail} | 跳过(断点): {skip}")
print(f"  有效结果: {len(valid_results)} 条")
print(f"  跳过(无预警): {len(skip_results)} 条")
print(f"  🔴 红色预警: {red_count} 条")
print(f"  🟡 黄色预警: {yellow_count} 条")
print(f"  Token 消耗: prompt={usage['prompt_tokens']} completion={usage['completion_tokens']} total={usage['total_tokens']}")
print(f"  输出文件: {OUTPUT_PATH}")
print("=" * 60)
